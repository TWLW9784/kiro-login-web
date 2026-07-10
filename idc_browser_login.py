# -*- coding: utf-8 -*-
"""idc_browser_login.py — Tu dong login trang device-code AWS IAM Identity Center
bang Playwright (Chromium SACH: khong cache / cookie / profile).

Luong (da soi DOM that tu d-9066713dd7.awsapps.com):
    1. Username page  : input[type=text] (#awsui-input-0)        -> "Next"
    2. Password page  : input[type=password] (#awsui-input-1)    -> "Sign in"
    3. (LAN DAU) Doi mat khau: 2 o password (new + confirm)      -> "Confirm"
    4. Authorization requested (hien user_code)                  -> "Confirm and continue"
    5. Allow kiro-oauth-client to access your data?              -> "Allow access"
    -> device-code da APPROVE => device_code_auth.poll_for_token() lay token durable.

Tool tu PHAT HIEN co form doi mat khau hay khong (qua so o password hien),
nen xu ly duoc ca 2 dang: lan dau (can doi pass) va lan sau (khong can).
"""
from __future__ import annotations

import math
import random
import re
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Callable, List, Optional
from urllib.parse import urlparse

from playwright.sync_api import sync_playwright

import mfa_totp


# Mac dinh
DEFAULT_IDC_START_URL = "https://d-9066713dd7.awsapps.com/start"
DEFAULT_NEW_PASSWORD = "Kiro@Durable2026#"   # du manh cho policy AWS (>=8, hoa/thuong/so/ky tu)

# Text marker (lowercase) de nhan dien trang thai
INCORRECT_MARKERS = (
    "incorrect", "is not correct", "could not sign", "couldn't sign",
    "authentication failed", "username or password", "try again later",
    "we couldn't", "no account", "does not exist", "invalid password",
)
MFA_MARKERS = (
    "multi-factor", "mfa device", "authenticator app", "verification code",
    "register mfa", "one-time passcode", "security key", "passkey",
)
# 「注册/绑定新 MFA」页特有特征：页面会展示二维码 / secret key / 引导扫码或手动录入。
# 只有出现这些才说明这是首次绑定页、页面上真有 AWS 生成的密钥可抠。
MFA_REGISTER_MARKERS = (
    "show secret key", "secret key", "scan the qr", "scan this qr", "qr code",
    "can't scan", "cannot scan", "enter key manually", "set up authenticator",
    "register mfa", "add mfa", "register your", "register a new", "assign mfa",
    "显示密钥", "显示秘钥", "扫描二维码", "设置身份验证", "注册", "绑定新",
)
# 「验证已有 MFA」页特有特征：账号已绑过 authenticator，AWS 只让输入 6 位现有验证码，
# 页面上没有任何可抠的密钥。这类页出现且无预设密钥时应立即失败、明确报错。
MFA_VERIFY_MARKERS = (
    "additional verification required", "enter the six digit code",
    "enter the 6 digit code", "six digit code from your authenticator",
    "code from your authenticator app", "troubleshoot mfa",
)
SUCCESS_MARKERS = (
    "request approved", "approved", "you can close", "added to your devices",
    "you may close", "request was approved",
)

# 微软联邦登录（Entra ID / Microsoft 365）相关。部分 AWS IDC 实例把身份
# 验证联邦给微软：设备码流程打开 start_url 后会重定向到 login.microsoftonline.com。
# 这时 AWS 原生 DOM（#awsui-input-*）对不上，需走微软的邮箱→密码→MFA 流程。
MS_LOGIN_HOSTS = (
    "login.microsoftonline.com", "login.microsoftonline.us", "login.partner.microsoftonline.cn",
    "login.live.com", "login.microsoft.com", "microsoftonline.com",
)
MS_PWD_ERROR_MARKERS = (
    "your account or password is incorrect", "incorrect password",
    "密码不正确", "couldn't find your account", "didn't recognize",
    "that microsoft account doesn't exist",
)


@dataclass
class LoginOutcome:
    ok: bool
    changed_password: bool = False
    error: str = ""
    mfa_secret: str = ""   # 若本次登录绑定了新 MFA，这里返回 AWS 生成的 TOTP 密钥（供回写保存）


@dataclass
class AccountRow:
    idx: int                 # so dong (xlsx: row 1-based; txt: line 1-based)
    email: str
    password: str
    proxy: str = ""
    kind: str = "xlsx"       # "xlsx" | "txt"


# ======================================================================
# Doc / ghi file account
# ======================================================================
_SEP_RE = re.compile(r"[\t,;|]")


def read_accounts_file(path: str | Path) -> List[AccountRow]:
    """Doc account tu .xlsx (Email|Password|Proxy) hoac .txt/.csv (email:password)."""
    p = Path(path)
    ext = p.suffix.lower()
    if ext in (".xlsx", ".xlsm"):
        return _read_xlsx(p)
    return _read_txt(p)


def _read_xlsx(p: Path) -> List[AccountRow]:
    from openpyxl import load_workbook
    wb = load_workbook(p)
    ws = wb.active
    out: List[AccountRow] = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row:
            continue
        email = _cell(row[0]) if len(row) > 0 else ""
        if not email or email.lower() in ("email", "username", "user", "account"):
            continue
        pw = _cell(row[1]) if len(row) > 1 else ""
        proxy = _cell(row[2]) if len(row) > 2 else ""
        out.append(AccountRow(idx=idx, email=email, password=pw, proxy=proxy, kind="xlsx"))
    wb.close()
    return out


def _read_txt(p: Path) -> List[AccountRow]:
    out: List[AccountRow] = []
    for i, raw in enumerate(p.read_text(encoding="utf-8", errors="replace").splitlines(), start=1):
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        # BO comment ket qua tool tu ghi-back ("   # OK/LOGIN FAIL ...").
        # Dung pattern ">=2 khoang trang + #" -> KHONG dung cham vao '#' trong password
        # (vd "Kiro@Durable2026#" giu nguyen vi khong co khoang trang truoc '#').
        line = re.split(r"\s{2,}#", line, maxsplit=1)[0].rstrip()
        if not line:
            continue
        # tach theo tab/comma/semicolon/pipe truoc, fallback dau ':'
        parts = _SEP_RE.split(line)
        if len(parts) < 2:
            parts = line.split(":", 2)  # email:pass[:proxy] (email khong chua ':')
        parts = [x.strip() for x in parts]
        
        # 识别 8 段管道格式（AWS IdC 导出）：
        # login_url | username | email | user_id | password | reset_at | status | error
        # AWS IdC 登录用 username（不带 @），不是 email
        if len(parts) >= 8 and ('awsapps.com' in parts[0] or 'amazonaws.com' in parts[0]):
            email = parts[1]  # username（AWS IdC 用 username 登录，不是 email）
            pw = parts[4]     # password
            proxy = ""        # 8 段格式不含 proxy
        else:
            # 原有格式：email|password|proxy 或 email:password:proxy
            email = parts[0] if parts else ""
            pw = parts[1] if len(parts) > 1 else ""
            proxy = parts[2] if len(parts) > 2 else ""
        
        # chap nhan username thuan (khong can @); can co password de tranh dong rac
        if not email or not pw:
            continue
        if email.lower() in ("email", "username", "user", "account"):
            continue
        out.append(AccountRow(idx=i, email=email, password=pw, proxy=proxy, kind="txt"))
    return out


def write_account_result(
    path: str | Path,
    acc: AccountRow,
    new_password: Optional[str],
    result: str,
) -> None:
    """Ghi ket qua + (neu doi pass) cap nhat password moi vao file de lan sau dung dung."""
    p = Path(path)
    try:
        if acc.kind == "xlsx":
            from openpyxl import load_workbook
            wb = load_workbook(p)
            ws = wb.active
            if new_password:
                ws.cell(row=acc.idx, column=2, value=new_password)   # cot B = Password
            ws.cell(row=acc.idx, column=4, value=result)             # cot D = Result
            wb.save(p)
            wb.close()
        else:
            # txt: rewrite dung dong, cap nhat password neu doi
            lines = p.read_text(encoding="utf-8", errors="replace").splitlines()
            if 1 <= acc.idx <= len(lines):
                if new_password:
                    sep = ":"
                    m = _SEP_RE.search(lines[acc.idx - 1])
                    if m:
                        sep = m.group(0)
                    pieces = [acc.email, new_password]
                    if acc.proxy:
                        pieces.append(acc.proxy)
                    lines[acc.idx - 1] = sep.join(pieces) + f"    # {result}"
                else:
                    base = lines[acc.idx - 1].split("    #", 1)[0].rstrip()
                    lines[acc.idx - 1] = base + f"    # {result}"
                p.write_text("\n".join(lines) + "\n", encoding="utf-8")
    except Exception:
        pass  # ghi ket qua that bai -> khong chan luong chinh


def _cell(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


# ======================================================================
# Playwright helpers
# ======================================================================

def _body_text(page) -> str:
    try:
        return (page.evaluate("() => document.body ? document.body.innerText : ''") or "")
    except Exception:
        return ""


def _count_visible(page, selector: str) -> int:
    try:
        return page.locator(f"{selector} >> visible=true").count()
    except Exception:
        try:
            return page.locator(selector).count()
        except Exception:
            return 0


def _maybe_cookie_accept(page, log) -> None:
    """Dong banner cookie AWS neu dang chan."""
    for sel in (
        'button[aria-label="Accept all cookies"]',
        'button[data-id="awsccc-cb-btn-accept"]',
    ):
        try:
            b = page.locator(sel).first
            if b.count() and b.is_visible():
                b.click(timeout=1500)
                return
        except Exception:
            pass


def _click_button(page, texts, log, timeout=2500) -> bool:
    """Bam button hien thi co text khop (exact uu tien, fallback contains)."""
    for t in texts:
        makers = (
            page.get_by_role("button", name=t, exact=True),
            page.get_by_role("button", name=t),
            page.locator(f'button:has-text("{t}")'),
        )
        for mk in makers:
            try:
                b = mk.first
                if b.count() and b.is_visible():
                    b.click(timeout=timeout)
                    log(f"    -> click '{t}'")
                    return True
            except Exception:
                pass
    try:
        btns = page.locator("button:visible, input[type=submit]:visible, [role=button]:visible")
        for i in range(min(btns.count(), 12)):
            b = btns.nth(i)
            try:
                label = " ".join(((b.inner_text(timeout=500) or b.get_attribute("value") or b.get_attribute("aria-label") or "").split()))
                if label and any(t.lower() in label.lower() for t in texts):
                    b.click(timeout=timeout)
                    log(f"    -> click '{label}'")
                    return True
            except Exception:
                pass
    except Exception:
        pass
    return False


def _fill_first_visible(page, selector: str, value: str, timeout=15000) -> bool:
    """Dien + VERIFY gia tri da vao field (chong race khi chay da luong)."""
    try:
        loc = page.locator(f"{selector} >> visible=true").first
        loc.wait_for(state="visible", timeout=timeout)
        loc.click(timeout=4000)
        try:
            loc.fill("")
        except Exception:
            pass
        loc.fill(value)
        # verify: cho chac field nhan dung gia tri truoc khi submit
        for _ in range(2):
            try:
                if (loc.input_value() or "") == value:
                    break
            except Exception:
                break
            try:
                loc.fill(value)
            except Exception:
                pass
            time.sleep(0.1)
        time.sleep(0.25)  # settle cho SPA dang ky onChange
        return True
    except Exception:
        return False


def _fill_nth_password(page, idx: int, value: str) -> None:
    """Dien o password thu idx (cho form doi mat khau), co verify."""
    try:
        loc = page.locator("input[type=password] >> visible=true").nth(idx)
        loc.click(timeout=3000)
        loc.fill(value)
        for _ in range(2):
            try:
                if (loc.input_value() or "") == value:
                    break
            except Exception:
                break
            loc.fill(value)
            time.sleep(0.1)
    except Exception:
        pass


def _visible_error_text(page) -> str:
    """Lay text loi THAT tu element alert (de log ly do that bai). '' neu khong co."""
    sel = ("[role=alert], [aria-live='assertive'], [aria-live='polite'], "
           "[class*='rror']")
    try:
        els = page.locator(sel)
        for i in range(min(els.count(), 8)):
            e = els.nth(i)
            try:
                if e.is_visible():
                    t = " ".join((e.inner_text() or "").split())
                    if len(t) >= 4:
                        return t[:140]
            except Exception:
                pass
    except Exception:
        pass
    return ""


def _wait_metadata1(page, timeout: float = 15.0, min_len: int = 500) -> bool:
    """Cho field an metadata1 (token chong-bot cua AWS, do JS sinh sau ~1.5s+) day du
    truoc khi submit. Quan trong khi chay DA LUONG (CPU cham -> JS sinh token lau hon).
    Tra True khi san sang (hoac khong co field do). False neu het timeout."""
    end = time.time() + timeout
    while time.time() < end:
        try:
            ln = page.evaluate(
                "() => { const el = document.querySelector('input[name=metadata1]');"
                " return el ? (el.value || '').length : -1; }")
        except Exception:
            ln = -1
        if ln == -1:      # khong co field -> khong can cho
            return True
        if ln >= min_len:
            return True
        time.sleep(0.25)
    return False


def _wait_password_resolve(page, timeout: float = 10.0) -> str:
    """Cho trang password chuyen tiep. TIN HIEU CAU TRUC (khong dua vao text body):
    'gone'  = da roi trang password (so o password != 1, hoac hien nut consent) -> OK,
    'stuck' = sau timeout van ket o trang password (sai pass / chua submit).
    """
    end = time.time() + timeout
    while time.time() < end:
        if _count_visible(page, "input[type=password]") != 1:
            return "gone"   # 0 = qua consent; >=2 = sang form doi mat khau
        if (_is_button_visible(page, "Allow access")
                or _is_button_visible(page, "Confirm and continue")):
            return "gone"
        time.sleep(0.3)
    return "stuck"


# ======================================================================
# Drive login
# ======================================================================

def _tile_args(window_index: int, window_count: int,
               screen_w: int, screen_h: int) -> list:
    """Tinh --window-position/--window-size de xep browser thanh luoi (headed)."""
    if window_count <= 1:
        return ["--start-maximized"]
    cols = min(4, max(1, math.ceil(math.sqrt(window_count))))
    rows = max(1, math.ceil(window_count / cols))
    w = max(560, screen_w // cols)
    h = max(520, screen_h // rows)
    slot = window_index % (cols * rows)
    x = (slot % cols) * w
    y = (slot // cols) * h
    return [f"--window-position={x},{y}", f"--window-size={w},{h}"]


def drive_login(
    verification_uri_complete: str,
    email: str,
    password: str,
    new_password: str = DEFAULT_NEW_PASSWORD,
    log: Callable[[str], None] = print,
    headless: bool = False,
    stop_event=None,
    proxy: str = "",
    timeout_s: int = 220,
    window_index: int = 0,
    window_count: int = 1,
    screen_w: int = 1920,
    screen_h: int = 1040,
    debug_dir: str = "",
    mfa_secret: str = "",
    on_secret=None,
    on_password_changed=None,
) -> LoginOutcome:
    """Mo Chromium sach, tu login trang device-code den khi bam 'Allow access'.

    on_secret(secret): 抠到 AWS 新生成的 MFA 密钥的瞬间立刻回调（完整、不脱敏），
    供上层立刻落盘，避免中途取消/超时导致密钥永久丢失。
    on_password_changed(new_pw): 微软/AWS 首次登录改密**提交成功的瞬间**立刻回调，
    供上层立刻落盘。防止改密后 Stay-signed-in/跳回 AWS/取 token 任一步超时或抛异常
    导致新密码永久丢失（历史 bug：落盘只在 drive_login 正常返回后靠 outcome.changed_password
    触发，改密后超时/异常返回会丢掉新密码 → 死号）。
    """
    launch_args = [
        "--disable-blink-features=AutomationControlled",
        "--disable-dev-shm-usage",
        "--disable-background-networking",
        "--disable-renderer-backgrounding",
        # 以下均为省 CPU/内存的标准 headless 参数：高并发时把 CPU 从
        # 渲染/GPU/后台任务上省下来，腾给反爬 token 的 JS 计算（实测
        # 4 核跑 10 并发时 token 生成被拖慢近 3×）。不禁 JS、不影响反爬检测。
        "--disable-gpu",
        "--disable-extensions",
        "--disable-background-timer-throttling",
        "--disable-hang-monitor",
        "--disable-features=TranslateUI,BackForwardCache",
        "--metrics-recording-only",
        "--no-first-run",
        "--mute-audio",
    ]
    if not headless:
        launch_args += _tile_args(window_index, window_count, screen_w, screen_h)
    launch_kwargs = {"headless": headless, "args": launch_args}
    if proxy:
        srv = proxy if "://" in proxy else f"http://{proxy}"
        launch_kwargs["proxy"] = {"server": srv}

    with sync_playwright() as pw:
        browser = pw.chromium.launch(**launch_kwargs)
        ctx_kwargs = {}
        if headless:
            ctx_kwargs["viewport"] = {"width": 1280, "height": 900}
        else:
            ctx_kwargs["no_viewport"] = True
        ctx = browser.new_context(**ctx_kwargs)
        try:
            ctx.route("**/*", lambda route: route.abort() if route.request.resource_type in ("image", "media", "font") else route.continue_())
        except Exception:
            pass
        page = ctx.new_page()
        page.set_default_timeout(15000)
        try:
            return _flow(page, verification_uri_complete, email, password,
                         new_password, log, stop_event, timeout_s, debug_dir,
                         known_mfa_secret=mfa_secret, on_secret=on_secret,
                         on_password_changed=on_password_changed)
        except Exception as e:
            return LoginOutcome(False, error=f"loi browser: {e}")
        finally:
            try:
                browser.close()
            except Exception:
                pass


# 验证码被 AWS 拒绝时页面常见的错误文案
MFA_REJECT_MARKERS = (
    "incorrect", "invalid", "not valid", "isn't valid", "does not match",
    "doesn't match", "wrong", "try again", "expired", "无效", "不正确",
    "错误", "不匹配", "重新",
)


def _handle_mfa_registration(page, log, _shot=None, known_secret: str = "",
                             on_secret=None) -> tuple[bool, str]:
    """处理 AWS「绑定新 MFA」流程（身份验证程序 / authenticator app）。

    步骤：
      1. 若页面让选择 MFA 类型，点「身份验证程序 / Authenticator app」
      2. 从页面抠出 AWS 生成的 secret key（mfa_totp.extract_secret_from_page）
      3. 本地算 TOTP 验证码填入（有些页面要求连续两个验证码）
      4. 提交后**验证 AWS 是否真的接受**（离开绑定页 / 出现 Allow access / success）
    返回 (ok, secret)：ok=是否确认绑定成功；secret=AWS 生成的密钥（供回写保存）。

    关键修复：旧版只要填了码、点了按钮就 `return True`，从不校验结果。
    若 AWS 因验证码错误/页面未真正提交而仍停在 MFA 页，主循环会因 mfa_secret
    已设而永久跳过 MFA 分支，表现为「卡在等待 Allow access 直到超时」。
    现在：提交后必须确认离开绑定页才算成功，否则换下一个 TOTP 窗口重试。

    密钥保全：抠到密钥的**瞬间**就通过 on_secret(secret) 回调落盘（完整、不脱敏），
    不等绑定结果。即使后续被取消/超时/AWS 卡住，密钥也已安全保存，绝不再丢失。
    """
    import time as _t

    def shot(tag):
        if _shot:
            try:
                _shot(tag)
            except Exception:
                pass

    def code_inputs():
        return page.locator(
            "input[type=text]:visible:not([readonly]):not([disabled]), "
            "input[type=tel]:visible:not([readonly]):not([disabled]), "
            "input[name*=code i]:visible, input[name*=otp i]:visible, "
            "input[name*=mfa i]:visible"
        )

    def fresh_totp() -> str:
        # 避免在 TOTP 周期末尾提交，AWS 页面/网络稍慢就会过期。
        remaining = 30 - (int(_t.time()) % 30)
        if remaining <= 8:
            log(f"    MFA: 当前验证码即将过期，等 {remaining + 1}s 换新码")
            _t.sleep(remaining + 1)
        return mfa_totp.totp_now(secret)

    def fill_code(index: int, value: str) -> bool:
        try:
            loc = code_inputs().nth(index)
            loc.click(timeout=2000)
            loc.fill("")
            loc.type(value, delay=20)
            try:
                return (loc.input_value(timeout=500) or "").strip() == value
            except Exception:
                return True
        except Exception:
            return False

    def submit_by_enter(input_count: int) -> bool:
        try:
            code_inputs().nth(max(0, min(input_count - 1, 1))).press("Enter")
            return True
        except Exception:
            try:
                page.keyboard.press("Enter")
                return True
            except Exception:
                return False

    def still_on_mfa() -> bool:
        try:
            b = _body_text(page).lower()
        except Exception:
            return True
        if any(s in b for s in SUCCESS_MARKERS):
            return False
        return any(m in b for m in MFA_MARKERS)

    def left_mfa_ok() -> bool:
        # 离开绑定页 / 出现授权页 / success 标记 = 绑定被接受
        try:
            b = _body_text(page).lower()
        except Exception:
            b = ""
        if any(s in b for s in SUCCESS_MARKERS):
            return True
        if _is_button_visible(page, "Allow access") or _is_button_visible(page, "Confirm and continue"):
            return True
        return not any(m in b for m in MFA_MARKERS)

    secret = known_secret or ""
    if not secret:
        # 1) 选择 authenticator app 类型（若有选项页）。重试时已在验证码页，跳过。
        for label in ("Authenticator app", "authenticator app", "身份验证程序",
                      "Authenticator", "Virtual", "TOTP"):
            try:
                loc = page.get_by_text(label, exact=False)
                if loc.count() > 0:
                    loc.first.click(timeout=1500)
                    log(f"    MFA: 选择「{label}」")
                    page.wait_for_timeout(500)
                    break
            except Exception:
                pass
        # 可能需要点 Next / Continue 进入密钥展示页
        _click_button(page, ["Next", "Continue", "下一步", "继续"], log, timeout=1500)
        page.wait_for_timeout(400)
        shot("mfa_setup_page")

        # 2) 抠 secret key
        secret = mfa_totp.extract_secret_from_page(page, log)
        if not secret:
            # 抠不到密钥：区分「注册页渲染慢/DOM 变了」与「其实是已有 MFA 的验证页」。
            # 若页面带验证页特征且无注册特征，则是已绑定账号，无密钥可抠，交由主循环秒退。
            try:
                _b = _body_text(page).lower()
            except Exception:
                _b = ""
            if any(m in _b for m in MFA_VERIFY_MARKERS) and not any(m in _b for m in MFA_REGISTER_MARKERS):
                log("    MFA: 当前为已有 authenticator 的验证页，无密钥可抠（账号已绑定 MFA）")
            return False, ""
        log(f"    MFA: 获取密钥成功（{secret[:4]}…{secret[-4:]}），本地计算验证码")
        # 立刻完整落盘（不等绑定结果），防止取消/超时导致密钥永久丢失
        if on_secret:
            try:
                on_secret(secret)
            except Exception:
                pass
    else:
        log(f"    MFA: 复用已抓取密钥（{secret[:4]}…{secret[-4:]}），重试验证码")

    # 3) 填码 + 提交 + 校验结果，最多 4 轮；每轮换新的 TOTP 窗口，避免重复用同一个码
    for attempt in range(1, 5):
        try:
            n = code_inputs().count()
        except Exception:
            n = 0

        if n == 0:
            # 没有验证码输入框：可能已离开绑定页 = 成功
            if left_mfa_ok():
                log("    MFA: 已无验证码框且离开绑定页，视为绑定成功")
                return True, secret
            page.wait_for_timeout(700)
            continue

        code = fresh_totp()
        if n >= 2:
            # 两框同屏：第一个填当前码，第二个等下一周期再填
            fill_code(0, code)
            wait = 31 - (int(_t.time()) % 30)
            log(f"    MFA: 需第二个连续码，等 {wait}s 到下一周期")
            _t.sleep(min(wait, 32))
            code2 = mfa_totp.totp_now(secret)
            fill_code(1, code2)
        else:
            if not fill_code(0, code):
                return False, secret

        shot(f"mfa_code_filled_{attempt}")
        enter_submitted = submit_by_enter(n)
        if enter_submitted:
            page.wait_for_timeout(1200)
            if left_mfa_ok():
                log("    MFA: Enter 提交后已离开绑定页/出现授权页，绑定成功")
                return True, secret
        clicked = _click_button(page, ["Assign MFA", "Add MFA", "Register", "Confirm",
                                       "Verify", "Submit", "Submit code", "Verify code",
                                       "Continue", "Next", "Done", "Finish",
                                       "绑定", "确认", "提交", "继续", "完成"], log)
        if not clicked and not enter_submitted:
            log("    MFA: 未找到提交按钮，且 Enter 提交失败")
        page.wait_for_timeout(2500)
        shot(f"mfa_after_submit_{attempt}")

        # —— 校验：是否真的被接受 ——
        if left_mfa_ok():
            log("    MFA: 提交后已离开绑定页/出现授权页，绑定成功")
            return True, secret

        # 仍在绑定页：检查是否报错（验证码被拒）
        try:
            body_now = _body_text(page).lower()
        except Exception:
            body_now = ""
        err = ""
        try:
            err = _visible_error_text(page) or ""
        except Exception:
            err = ""
        rejected = bool(err) or any(m in body_now for m in MFA_REJECT_MARKERS)
        if rejected:
            log(f"    MFA: 验证码被拒（{err or '页面提示无效/重试'}），换下一窗口码重试")
        else:
            log("    MFA: 提交后仍停留绑定页，换下一窗口码重试")

        # 等到下一个 TOTP 窗口，避免重复使用刚才那个码
        wait = 31 - (int(_t.time()) % 30)
        _t.sleep(min(wait, 32))

    log("    MFA: 多轮提交后仍未离开绑定页，放弃本次绑定")
    return False, secret


def _on_ms_login(page) -> bool:
    """当前页是否已重定向到微软登录（Entra/M365）。"""
    try:
        host = (urlparse(page.url).hostname or "").lower()
    except Exception:
        return False
    return any(host == h or host.endswith("." + h) or host.endswith(h) for h in MS_LOGIN_HOSTS)


def _handle_ms_federated(page, email, password, log, _shot, stop_event,
                         mfa_secret: str = "", deadline: float = 0.0,
                         new_password: str = "", on_password_changed=None) -> tuple[str, str, bool]:
    """处理 AWS IDC 联邦到微软（Entra ID / Microsoft 365）的登录页。

    微软页 DOM 与 AWS 原生页完全不同，这里复用 m365 驱动的 selector：
      邮箱页（loginfmt/type=email）→ Next → 密码页（passwd）→ Sign in
      → [可选 首次登录改密 Update your password] → [可选 MFA] → Stay signed in? → Yes
    处理完会自动跳回 AWS 同意页（Allow access），交回主循环继续。

    返回 (status, detail, changed_password)：
      'left'  = 已离开微软页（跳回 AWS 或 success），交回主循环
      'error' = 硬错误（密码错/需 MFA 但无密钥等），detail 为原因
      'timeout' = 本阶段超时
    changed_password：本次是否在微软页成功改密（改密后新密码=new_password）。
    """
    mfa_norm = mfa_totp.normalize_secret(mfa_secret) if mfa_secret else ""
    if deadline <= 0:
        deadline = time.time() + 120
    log("    检测到联邦至微软登录页（Entra/M365），切换微软 DOM 流程")
    _shot("ms_login")
    changed_pw = False
    # 状态驱动：不用 sticky 的 email_done/pwd_done 锁（历史 bug：Next 未真正推进时
    # _pw_ready() 被预渲染的隐藏密码框骗出 false-positive → 往无效框填密码并焊死 pwd_done
    # → 页面仍在邮箱页 → 所有分支不匹配 → 空转到超时）。改为每轮按当前真实页面状态决策，
    # 配合各步尝试次数上限 + 未识别页面心跳日志/截图，杜绝黑盒死等。
    email_attempts = 0
    pwd_attempts = 0
    change_pw_attempts = 0
    mfa_attempts = 0
    last_url = ""
    idle = 0
    last_hb = 0.0

    def _pw_ready() -> bool:
        # 微软邮箱→密码有导航，passwd 框真正可交互（可见+有尺寸+未禁用）才算到位
        try:
            return page.evaluate(
                "()=>{const el=document.querySelector('input[name=passwd]');"
                "if(!el)return false;const cs=getComputedStyle(el);"
                "const r=el.getBoundingClientRect();"
                "return cs.display!=='none'&&cs.visibility!=='hidden'&&r.height>0"
                "&&!el.disabled&&!el.readOnly;}")
        except Exception:
            return False

    def _email_ready() -> bool:
        # 邮箱框真正可交互（可见+有尺寸+未禁用）才算还停在账号输入页
        try:
            return page.evaluate(
                "()=>{const el=document.querySelector('input[type=email],input[name=loginfmt],input[name=email]');"
                "if(!el)return false;const cs=getComputedStyle(el);"
                "const r=el.getBoundingClientRect();"
                "return cs.display!=='none'&&cs.visibility!=='hidden'&&r.height>0"
                "&&!el.disabled&&!el.readOnly;}")
        except Exception:
            return False

    def _advanced_from_email(timeout: float = 14.0) -> bool:
        # 填完邮箱点 Next 后，确认页面真的推进：密码框就绪 / 出现账号错误 / 离开微软登录域
        # 历史抑制阵——不能因为 `not _email_ready()` 就秒返 True，邮箱框在导航中会瞬时消失，
        # 一旦重新出现又会被主循环误判回到邮箱分支 → 一直重填邮箱、却向已推进的密码页提交空密码。
        end = time.time() + timeout
        while time.time() < end:
            time.sleep(0.5)
            if _pw_ready():
                return True
            try:
                b = page.inner_text("body", timeout=1500).lower()
            except Exception:
                b = ""
            if any(k in b for k in MS_PWD_ERROR_MARKERS):
                return True
            if not _on_ms_login(page):
                return True
        return False

    while time.time() < deadline:
        if stop_event is not None and stop_event.is_set():
            return "error", "已中断", changed_pw
        # 已离开微软页 = 跳回 AWS（同意页/success），交回主循环
        if not _on_ms_login(page):
            log("    微软登录完成，已跳回 AWS，交回主循环处理同意页")
            _shot("ms_left_to_aws")
            return "left", "", changed_pw
        try:
            body = page.inner_text("body", timeout=3000).lower()
        except Exception:
            body = ""

        # 1) 首次登录/密码过期：Update your password（current + new + confirm 三框）——最具体，先判
        cur_box = page.locator("input[name=currentpasswd], #currentPassword").first
        new_box = page.locator("input[name=newpasswd], #newPassword").first
        if (("update your password" in body or "更新密码" in body
                or (cur_box.count() and new_box.count()))
                and cur_box.count() and new_box.count()):
            change_pw_attempts += 1
            if change_pw_attempts > 3:
                _shot("ms_change_pw_stuck")
                detail = _visible_error_text(page)
                return "error", f"微软首次登录改密多次未通过（{detail or '密码策略不符/页面未推进'}）", changed_pw
            eff_new = new_password or password
            try:
                cur_box.fill(password)
                new_box.fill(eff_new)
                page.locator("input[name=confirmnewpasswd], #confirmNewPassword").first.fill(eff_new)
                log(f"    微软首次登录改密：已填入当前/新/确认密码（新密码={eff_new}）")
                _shot("ms_update_pw")
                _click_ms(page, ['#idSIButton9', 'input[type=submit]',
                                 'button[type=submit]', 'button:has-text("Sign in")'])
                if eff_new != password:
                    changed_pw = True
                    # 改密提交成功的瞬间立刻落盘（微软联邦改密路径），防止后续
                    # Stay-signed-in/跳回 AWS/取 token 超时或异常丢新密码 → 死号。
                    if on_password_changed and eff_new:
                        try:
                            on_password_changed(eff_new)
                        except Exception:
                            pass
                time.sleep(3.0)
                continue
            except Exception:
                time.sleep(1.0)
                continue

        # 2) 密码错 / 账号不存在
        if any(k in body for k in MS_PWD_ERROR_MARKERS):
            _shot("ms_pwd_error")
            detail = _visible_error_text(page)
            return "error", f"微软账号或密码错误（{detail or 'account or password incorrect'}）", changed_pw

        # 3) MFA：需 TOTP 密钥才能自动
        if any(k in body for k in ("verification code", "enter code", "authenticator",
                                   "verify your identity", "输入验证码", "one-time",
                                   "approve sign in", "身份验证")):
            mfa_attempts += 1
            if mfa_attempts > 4:
                return "error", "微软 MFA 验证多次失败", changed_pw
            if mfa_norm:
                code = mfa_totp.totp_now(mfa_norm)
                code_box = page.locator("input[type=tel], input[name=otc], "
                                        "input[autocomplete='one-time-code'], input[type=text]").first
                if code_box.count() and code_box.is_visible():
                    try:
                        code_box.fill(code)
                        _click_ms(page, ['#idSubmit_SAOTCC_Continue', '#idSIButton9',
                                         'input[type=submit]', 'button:has-text("Verify")',
                                         'button:has-text("Next")'])
                        log(f"    已填入微软 MFA 验证码（第 {mfa_attempts} 次）")
                        _shot(f"ms_mfa_{mfa_attempts}")
                        time.sleep(3.0)
                        continue
                    except Exception:
                        pass
            else:
                _shot("ms_mfa_no_secret")
                return "error", "该账号联邦微软且要求 MFA，但未提供 TOTP 密钥（暂不支持自动绑定微软 authenticator）", changed_pw

        # 4) Stay signed in?
        if "stay signed in" in body or "保持登录" in body or "reduce the number of times" in body:
            _click_ms(page, ['#idSIButton9', 'button:has-text("Yes")',
                             'input[type=submit][value="Yes"]'])
            log("    Stay signed in → Yes")
            time.sleep(2.5)
            continue

        # 5) More information required / 跳过类提示
        if "more information" in body or "需要更多信息" in body:
            if not _click_ms(page, ['a:has-text("Skip")', 'button:has-text("Skip")',
                                    'input[value="Skip setup"]', '#idBtn_Back']):
                _click_ms(page, ['#idSIButton9', 'input[type=submit]'])
            time.sleep(2.5)
            continue

        # 6) 密码页：passwd 框真正可交互，且【已不在邮箱页】。
        #    加 not _email_ready() 守卫：堵住邮箱页预渲染隐藏密码框的 false-positive
        #    （历史死锁根因——Next 没推进却误填密码）。
        if _pw_ready() and not _email_ready():
            pwd_attempts += 1
            if pwd_attempts > 4:
                _shot("ms_pwd_stuck")
                return "error", "微软密码页多次提交未推进（密码错或页面加载慢）", changed_pw
            pw = page.locator("input[type=password], input[name=passwd]").first
            try:
                pw.fill(password)
                # 回读确认真的填进去了（避免向密码页提交空密码 → "Please enter your password."）
                try:
                    filled_val = pw.input_value()
                except Exception:
                    filled_val = ""
                if not filled_val:
                    log("    微软密码填入后回读为空，重试填入")
                    time.sleep(0.5)
                    try:
                        pw.click()
                    except Exception:
                        pass
                    pw.type(password, delay=30)
                    try:
                        filled_val = pw.input_value()
                    except Exception:
                        filled_val = ""
                if not filled_val:
                    _shot("ms_pwd_fill_empty")
                    time.sleep(1.0)
                    continue
                log("    已填入微软密码")
                _shot("ms_pwd_filled")
                _click_ms(page, ['#idSIButton9', 'input[type=submit]',
                                 'button[type=submit]', 'button:has-text("Sign in")',
                                 'button:has-text("登录")'])
                # 确认离开密码页（就绪消失）或出现错误，否则下一轮受 pwd_attempts 限制重试
                end = time.time() + 10.0
                while time.time() < end:
                    time.sleep(0.5)
                    if not _pw_ready():
                        break
                    try:
                        b2 = page.inner_text("body", timeout=1500).lower()
                    except Exception:
                        b2 = ""
                    if any(k in b2 for k in MS_PWD_ERROR_MARKERS):
                        break
                continue
            except Exception:
                time.sleep(1.0)
                continue

        # 7) 邮箱页：账号框可交互就填 + Next，并【验证真的推进】，没推进则重试（有上限）
        if _email_ready():
            email_attempts += 1
            if email_attempts > 4:
                _shot("ms_email_stuck")
                return "error", "微软邮箱页多次提交未推进（Next 未生效或页面加载慢）", changed_pw
            email_box = page.locator("input[type=email], input[name=loginfmt], input[name=email]").first
            try:
                email_box.fill(email)
                log(f"    已填入微软邮箱/账号（第 {email_attempts} 次）")
                _shot("ms_email_filled")
                _click_ms(page, ['#idSIButton9', 'input[type=submit]',
                                 'button[type=submit]', 'button:has-text("Next")',
                                 'button:has-text("下一步")'])
                if not _advanced_from_email():
                    log("    微软邮箱提交后页面未推进（Next 可能未生效），将重试")
                    _shot("ms_email_no_advance")
                continue
            except Exception:
                time.sleep(1.0)
                continue

        # 8) 未识别页面兜底：心跳日志 + 截图（每 12s 一次），杜绝黑盒；同 URL 停滞尝试通用前进
        url = page.url
        idle = idle + 1 if url == last_url else 0
        last_url = url
        now = time.time()
        if now - last_hb > 12:
            title = ""
            try:
                title = page.title()
            except Exception:
                pass
            log(f"    微软页等待中（未识别状态）：title={title!r} url={url[:90]}")
            _shot("ms_unknown")
            last_hb = now
        if idle >= 2:
            _click_ms(page, ['#idSIButton9', 'button[type=submit]', 'input[type=submit]',
                             'button:has-text("Next")', 'button:has-text("Continue")'])
        time.sleep(1.5)
    return "timeout", "微软登录阶段超时", changed_pw


def _click_ms(page, selectors) -> bool:
    for sel in selectors:
        try:
            el = page.locator(sel).first
            if el.count() and el.is_visible():
                el.click(timeout=2500)
                return True
        except Exception:
            continue
    return False


def _flow(page, url, email, password, new_password, log, stop_event, timeout_s,
          debug_dir: str = "", known_mfa_secret: str = "", on_secret=None,
          on_password_changed=None) -> LoginOutcome:
    safe = "".join(c if c.isalnum() else "_" for c in (email or "acc"))[:20]
    shot = {"i": 0}

    def _shot(tag: str):
        if not debug_dir:
            return
        try:
            from pathlib import Path as _P
            _P(debug_dir).mkdir(exist_ok=True, parents=True)
            page.screenshot(path=str(_P(debug_dir) / f"{safe}_{shot['i']:02d}_{tag}.png"))
        except Exception:
            pass
        shot["i"] += 1

    # 轻微错峰，避免高并发同时打同一个登录入口
    time.sleep(random.uniform(0.02, 0.12))
    log("    mo trang login (clean, no cache)...")
    page.goto(url, wait_until="domcontentloaded")
    try:
        page.wait_for_selector("input", timeout=12000)
    except Exception:
        pass

    deadline = time.time() + timeout_s
    changed = False
    pw_attempts = 0
    password_candidates = [password]
    if new_password and new_password != password:
        password_candidates.append(new_password)
    change_pw_attempts = 0
    user_attempts = 0
    idle_rounds = 0
    seen_password = False
    # 已绑定 MFA 的账号：用户提前提供 TOTP 密钥（base32），登录时直接用它算验证码。
    # 新账号首次绑定：留空，AWS 会展示密钥，由 _handle_mfa_registration 自动抠出来。
    mfa_secret = mfa_totp.normalize_secret(known_mfa_secret) if known_mfa_secret else ""
    if mfa_secret:
        log(f"    MFA: 已注入预设密钥（{mfa_secret[:4]}…{mfa_secret[-4:]}），将用于 TOTP 验证")
    mfa_attempts = 0

    while time.time() < deadline:
        if stop_event is not None and stop_event.is_set():
            return LoginOutcome(False, changed, "Da huy")

        _maybe_cookie_accept(page, log)
        body = _body_text(page).lower()

        # --- 联邦至微软（Entra/M365）：部分 AWS IDC 实例把身份验证交给微软。
        #     微软页 DOM 与 AWS 原生页完全不同，必须先于 AWS 分支检测并切换流程。---
        if _on_ms_login(page):
            seen_password = True  # 已进入联邦登录，防止回到 username 分支误判
            st_ms, detail_ms, changed_ms = _handle_ms_federated(
                page, email, password, log, _shot, stop_event,
                mfa_secret=mfa_secret, deadline=deadline, new_password=new_password,
                on_password_changed=on_password_changed,
            )
            if changed_ms:
                changed = True
            if st_ms == "left":
                # 已跳回 AWS（同意页/success），交回主循环继续处理 Allow access
                idle_rounds = 0
                time.sleep(0.8)
                continue
            if st_ms == "error":
                return LoginOutcome(False, changed, detail_ms, mfa_secret=mfa_secret)
            # timeout：回主循环（可能已接近总超时）
            continue

        # --- MFA 绑定：新账号首次被要求绑定 authenticator。
        #     自动抠密钥 + 本地算 TOTP 填入完成绑定，并保存密钥。
        #     修复：只要页面还在 MFA 绑定页就继续处理（不再用 not mfa_secret 守卫永久跳过），
        #     已抠到密钥的重试直接复用。---
        if any(m in body for m in MFA_MARKERS):
            mfa_attempts += 1
            # 区分「注册/绑定新 MFA」与「验证已有 MFA」。
            # 无预设密钥时：若页面是验证页（无注册特征，且带验证页特征），
            # 说明账号已绑过 authenticator、密钥不在我们手里，抠不出任何密钥。
            # 直接秒退，报准确错误（且不含风控/验证码关键词，避免误触发换 IP 重试）。
            if not mfa_secret:
                has_register = any(m in body for m in MFA_REGISTER_MARKERS)
                is_verify = any(m in body for m in MFA_VERIFY_MARKERS)
                if is_verify and not has_register:
                    log("    MFA: 页面为已有 authenticator 的验证页（非首次绑定），无可抠密钥")
                    if _shot:
                        _shot("mfa_already_bound")
                    return LoginOutcome(
                        False, changed,
                        "账号已绑定 MFA，需提供已有 authenticator 密钥（页面无密钥可抓）",
                        mfa_secret=mfa_secret,
                    )
            if mfa_attempts > 3:
                return LoginOutcome(
                    False, changed,
                    "MFA 绑定多次未完成（未抓到密钥或验证码被拒）",
                    mfa_secret=mfa_secret,
                )
            log("    检测到 MFA 绑定页，尝试自动绑定 authenticator")
            ok_mfa, sec = _handle_mfa_registration(page, log, _shot, known_secret=mfa_secret,
                                                   on_secret=on_secret)
            if sec:
                mfa_secret = sec
            if not ok_mfa:
                # 未抠到密钥或提交未被接受：稍等后主循环重新进入本分支重试
                time.sleep(1.0)
                continue
            idle_rounds = 0
            time.sleep(1.0)
            continue

        # --- da Allow xong / approved ---
        if any(m in body for m in SUCCESS_MARKERS):
            log("    -> approved")
            return LoginOutcome(True, changed, mfa_secret=mfa_secret)

        pw_n = _count_visible(page, "input[type=password]")
        text_n = _count_visible(page, "input[type=text]:not([type=hidden])")
        has_allow = _is_button_visible(page, "Allow access")
        has_confirm = _is_button_visible(page, "Confirm and continue")

        # --- consent: Allow access (cuoi cung) ---
        if has_allow:
            if _click_button(page, ["Allow access"], log):
                time.sleep(0.8)
                return LoginOutcome(True, changed, mfa_secret=mfa_secret)   # approve xong -> poll se lay token

        # --- consent: Confirm and continue (xac nhan user_code) ---
        if has_confirm:
            _click_button(page, ["Confirm and continue"], log)
            idle_rounds = 0
            time.sleep(0.8)
            continue

        # --- DOI MAT KHAU lan dau: >=2 o password ---
        if pw_n >= 2:
            change_pw_attempts += 1
            log(f"    form DOI MAT KHAU lan dau -> dat password moi（新密码={new_password}）")
            if change_pw_attempts > 3:
                detail = _visible_error_text(page)
                return LoginOutcome(
                    False,
                    changed,
                    f"首次登录改密码失败：{detail or '页面未进入下一步'}",
                )
            if pw_n >= 3:
                # current / new / confirm
                _fill_nth_password(page, 0, password)
                _fill_nth_password(page, 1, new_password)
                _fill_nth_password(page, 2, new_password)
            else:
                _fill_nth_password(page, 0, new_password)
                _fill_nth_password(page, 1, new_password)
            time.sleep(0.3)
            _wait_metadata1(page)
            clicked = _click_button(
                page,
                ["Confirm", "Change password", "Update password", "Set new password",
                 "Set password", "Save changes", "Save", "Continue", "Submit"],
                log,
            )
            if not clicked:
                try:
                    page.keyboard.press("Enter")
                except Exception:
                    pass
            idle_rounds = 0
            end = time.time() + 8.0
            while time.time() < end:
                body_after = _body_text(page).lower()
                if any(m in body_after for m in INCORRECT_MARKERS):
                    detail = _visible_error_text(page)
                    return LoginOutcome(False, changed, f"首次登录改密码失败：{detail or '密码策略不通过或登录被拒绝'}")
                if (_is_button_visible(page, "Allow access")
                        or _is_button_visible(page, "Confirm and continue")
                        or any(m in body_after for m in SUCCESS_MARKERS)):
                    break
                next_pw_n = _count_visible(page, "input[type=password]")
                if next_pw_n < 2:
                    break
                time.sleep(0.3)

            if _count_visible(page, "input[type=password]") >= 2:
                detail = _visible_error_text(page)
                log(f"    首次登录改密码未通过（{detail or '仍停留在改密码页'}）")
                continue

            changed = True
            # 必须确认 AWS 已接受新密码（离开改密页/出现授权页）后再落盘。
            # 旧逻辑在点击按钮后立即保存，遇到密码策略不通过时会把失败的弱密码写入，
            # 后续强密码重试又因 saved 标记无法覆盖，造成“强密码未保存”。
            if on_password_changed and new_password:
                try:
                    on_password_changed(new_password)
                except Exception:
                    pass
            time.sleep(0.5)
            continue

        # --- PASSWORD page: 1 o password ---
        if pw_n == 1:
            seen_password = True
            pw_attempts += 1
            candidate = password_candidates[min(pw_attempts - 1, len(password_candidates) - 1)]
            log(f"    nhap password (lan {pw_attempts})" + ("，尝试首次登录新密码" if candidate == new_password and candidate != password else ""))
            _wait_metadata1(page)            # cho token anti-bot truoc khi dien
            _fill_first_visible(page, "input[type=password]", candidate)
            _shot("pw_filled")
            if debug_dir:
                try:
                    val = page.locator("input[type=password] >> visible=true").first.input_value()
                    log(f"    [debug] o password dang chua: {val!r} (khop pass={val == candidate})")
                except Exception:
                    pass
            if not _wait_metadata1(page):
                log("    WARN: metadata1 chua san sang (van thu submit)")
            if not _click_button(page, ["Sign in", "Continue", "Next"], log):
                try:
                    page.keyboard.press("Enter")
                except Exception:
                    pass
            # roi trang password = OK; ket lai = sai pass / chua submit
            st = _wait_password_resolve(page, timeout=12.0)
            _shot(f"after_signin_{st}")
            if st == "gone":
                idle_rounds = 0
                continue
            detail = _visible_error_text(page)
            # captcha 专项识别：触发人机验证时，再试密码也没用（验证码会一直在），直接明确报错。
            if detail and any(m in detail.lower() for m in ("captcha", "验证码", "are you human", "robot")):
                _shot("captcha")
                return LoginOutcome(
                    False, changed,
                    f"触发 AWS 人机验证 captcha（{detail}）：请降低并发至 1~2 并挂代理（不同 IP）后重试")
            if pw_attempts >= len(password_candidates):
                return LoginOutcome(
                    False, changed,
                    f"Login that bai o trang password: {detail or 'khong qua duoc (sai pass?)'}")
            log(f"    chua qua password ({detail or 'thu lai'}) -> 换下一个候选密码重试")
            idle_rounds = 0
            continue

        # --- USERNAME page: co text input, chua co password ---
        if text_n >= 1 and pw_n == 0 and not has_allow and not has_confirm:
            # Da qua password roi ma quay lai username = sign-in bi tu choi/reset
            if seen_password:
                detail = _visible_error_text(page)
                _shot("reset_to_username")
                return LoginOutcome(
                    False, changed,
                    f"Sign-in bi tu choi (reset ve username): {detail or 'anti-bot/sai pass'}")
            user_attempts += 1
            if user_attempts > 3:
                return LoginOutcome(False, changed, "Khong qua duoc buoc username (email sai?).")
            log("    nhap username/email")
            _wait_metadata1(page)
            _fill_first_visible(page, "#awsui-input-0, input[type=text]", email)
            _shot("user_filled")
            if not _wait_metadata1(page):
                log("    WARN: metadata1 chua san sang (van thu submit)")
            if not _click_button(page, ["Next", "Continue", "Sign in"], log):
                try:
                    page.keyboard.press("Enter")
                except Exception:
                    pass
            # cho chuyen sang trang password (tranh re-fire username khi parallel cham)
            try:
                page.wait_for_selector("input[type=password] >> visible=true", timeout=8000)
            except Exception:
                pass
            idle_rounds = 0
            continue

        # --- khong co gi de lam ---
        idle_rounds += 1
        # Da qua password nhung chua thay nut consent/success: khong duoc coi nhu approved.
        # Neu o day tra ok, buoc CreateToken se doi authorization_pending den het han,
        # nhin nhu "卡住没有日志"。改为持续等待并周期性提示。
        if pw_attempts >= 1 and idle_rounds in (4, 10, 20):
            log("    da qua password, dang doi trang xac nhan/Allow access")
        time.sleep(0.6)

    return LoginOutcome(False, changed, "Timeout - khong hoan tat login trong thoi gian cho.")


def _is_button_visible(page, text: str) -> bool:
    try:
        b = page.get_by_role("button", name=text, exact=True).first
        if b.count() and b.is_visible():
            return True
    except Exception:
        pass
    try:
        b = page.locator(f'button:has-text("{text}")').first
        return bool(b.count() and b.is_visible())
    except Exception:
        return False


__all__ = [
    "LoginOutcome", "AccountRow",
    "read_accounts_file", "write_account_result",
    "drive_login", "DEFAULT_IDC_START_URL", "DEFAULT_NEW_PASSWORD",
]
